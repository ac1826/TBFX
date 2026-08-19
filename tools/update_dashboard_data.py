from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


REPO_ROOT = Path(__file__).resolve().parents[1]
STATIC_DIR = REPO_ROOT / "streamlit_html_portal" / "static"
SALES_STATIC = STATIC_DIR / "2026-sales-dashboard.html"
YOY_STATIC = STATIC_DIR / "2026-vs-2025-yoy-dashboard.html"

F_DIR = Path("F:/llqdocument/大成文件/客户贡献分析")
SOURCE_2026_NAME = "PFS 26年毛利表1-7月 含电商 TOP20新.xlsx"
SOURCE_2025_NAME = "PFS 25年毛利表1-7月 含电商 TOP20 -品项实际 新.xlsx"
YOY_MONTH_ORDER = [f"{month}月" for month in range(1, 8)]
YOY_CHANNEL_ORDER = ["烘焙", "休闲", "团膳", "宴席", "零售", "KA", "鲜食工厂", "其他"]
CUSTOMER_MASTER_NAME = "7月客户渠道及合并汇总.xlsx"
CHANNEL_CLASSIFICATION_NAME = "渠道分类.xlsx"
ORG_MASTER_NAME = "销售办事处组织架构及渠道.xlsx"
CUSTOMER_LIST_NAME = "客户清单.xlsx"
REFERENCE_HTML_NAME = "2026年1-6月主次渠道分析仪表盘.html"
SALES_SOURCE_HTML = F_DIR / "2026年1-7月数据分析仪表盘.html"
YOY_SOURCE_HTML = F_DIR / "26年与25年_1-7月数据对比分析.html"

BAD_CODES = {"CA2428001", "CC1131011"}
DELETE_MARKER = "删"
UNMAPPED = "未映射"
UNMAPPED_OUTPUT = F_DIR / "未映射明细.xlsx"

# Explicit location supplements confirmed from public company/government records.
# These are keyed by customer ID; office names are never used to infer geography.
RESEARCHED_GEO: dict[str, dict[str, str]] = {
    "GBR0000": {"p": "香港特别行政区", "ct": "中国香港", "source": "用户明确规则"},
    "274055": {"p": "广东省", "ct": "东莞", "source": "https://www.evirthfood.com/news.html"},
    "GW15": {"p": "辽宁省", "ct": "大连", "source": "https://www.dfa3999.com/sc/responsibility_news_details.php?id=36"},
    "299149": {"p": "广东省", "ct": "广州", "source": "https://www.gz.gov.cn/attachment/7/7972/7972121/10684809.pdf"},
    "262906": {"p": "河南省", "ct": "南阳", "source": "客户完整工商名称（南阳市天禄商贸有限公司）"},
    "262907": {"p": "江西省", "source": "客户完整工商名称（江西特源贸易有限公司）"},
    "262908": {"p": "河南省", "ct": "漯河", "source": "客户完整工商名称（漯河元粒食品有限公司）"},
    "274620": {"p": "江苏省", "ct": "昆山", "source": "https://m.zhipin.com/companys/061ebf4222b1ee0f03x73tW-GVI~.html"},
    "274732": {"p": "甘肃省", "ct": "庆阳", "source": "https://www.shuididp.cn/yp-21fa6738f82d709caa3d2daf0fb5a001.html"},
    "274740": {"p": "云南省", "source": "客户完整工商名称（云南吉满福供应链管理有限公司）"},
    "275929": {"p": "广东省", "source": "客户完整工商名称（广东唐馆食品有限公司）"},
    "275930": {"p": "海南省", "source": "客户完整工商名称（海南蓝天浩瀚食品有限公司）"},
    "275932": {"p": "广东省", "ct": "汕尾", "source": "客户完整工商名称（陆丰市甲子镇炜灿食品）"},
    "275933": {"p": "广东省", "ct": "佛山", "source": "https://www.qiyeku.cn/b2b/4705594.html"},
    "285002": {"p": "山东省", "ct": "德州", "source": "https://static1.tianyancha.com/companyHonorLabel/gov-notice/file/edaeb37bc25ae80fccf6b6bbdad0ea87.pdf"},
}


def find_source_file(name: str) -> Path:
    for root in (F_DIR, Path("E:/")):
        if not root.exists():
            continue
        for path in root.rglob(name):
            if path.is_file():
                return path
    raise FileNotFoundError(f"Cannot find source workbook: {name}")


def copy_sources_to_f_dir() -> tuple[Path, Path]:
    F_DIR.mkdir(parents=True, exist_ok=True)
    src_2026 = find_source_file(SOURCE_2026_NAME)
    src_2025 = find_source_file(SOURCE_2025_NAME)
    dst_2026 = F_DIR / SOURCE_2026_NAME
    dst_2025 = F_DIR / SOURCE_2025_NAME
    if src_2026.resolve() != dst_2026.resolve():
        shutil.copy2(src_2026, dst_2026)
    if src_2025.resolve() != dst_2025.resolve():
        shutil.copy2(src_2025, dst_2025)
    return dst_2026, dst_2025


def extract_app_data(text: str) -> tuple[int, int, dict[str, Any]]:
    match = re.search(r"(?:const|let|var)\s+APP_DATA\s*=\s*", text)
    if not match:
        raise ValueError("APP_DATA assignment not found")

    start = match.end()
    while start < len(text) and text[start].isspace():
        start += 1
    if start >= len(text) or text[start] not in "{[":
        raise ValueError("APP_DATA does not start with an object or array")

    stack: list[str] = []
    in_string = False
    quote = ""
    escaped = False
    for pos in range(start, len(text)):
        char = text[pos]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == quote:
                in_string = False
            continue
        if char in ('"', "'"):
            in_string = True
            quote = char
        elif char in "[{":
            stack.append("}" if char == "{" else "]")
        elif char in "]}":
            if not stack or char != stack[-1]:
                raise ValueError("Unbalanced APP_DATA JSON")
            stack.pop()
            if not stack:
                raw = text[start : pos + 1]
                return start, pos + 1, json.loads(raw)
    raise ValueError("APP_DATA JSON end not found")


def replace_app_data(path: Path, data: dict[str, Any]) -> None:
    text = path.read_text(encoding="utf-8", errors="replace")
    start, end, _ = extract_app_data(text)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"), allow_nan=False)
    path.write_text(text[:start] + payload + text[end:], encoding="utf-8")


def load_app_data(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    _, _, data = extract_app_data(text)
    return data


def clean_str(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)) or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_code(value: Any) -> str:
    text = clean_str(value)
    if text.endswith(".0"):
        return text[:-2]
    return text


def num(value: Any) -> float:
    if value is None or pd.isna(value):
        return 0.0
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return 0.0
    return float(parsed)


def k(value: float, digits: int = 6) -> float:
    result = value / 1000.0
    if abs(result) < 0.0000005:
        return 0.0
    return round(result, digits)


def month_2026(value: Any) -> tuple[str, str]:
    raw = clean_code(value)
    if not raw:
        return "", ""
    if raw.startswith("26") and len(raw) >= 4:
        month = int(raw[-2:])
        return f"2026.{month:02d}", f"26{month:02d}"
    month = int(float(raw))
    return f"2026.{month:02d}", f"26{month:02d}"


def month_cn(value: Any, year: str) -> str:
    raw = clean_str(value).removesuffix("月").strip()
    if not raw:
        return ""
    code = clean_code(raw)
    if code.startswith(year) and len(code) >= 6:
        month = int(code[-2:])
    elif code.startswith(year[-2:]) and len(code) >= 4:
        month = int(code[-2:])
    else:
        month = int(float(code))
    if not 1 <= month <= 12:
        raise ValueError(f"Invalid month value for {year}: {value}")
    return f"{month}月"


def most_common_lookup(records: list[dict[str, Any]], key: str, fields: tuple[str, ...]) -> dict[str, dict[str, str]]:
    buckets: dict[str, Counter[tuple[str, ...]]] = defaultdict(Counter)
    for record in records:
        key_value = clean_code(record.get(key))
        if not key_value:
            continue
        values = tuple(clean_str(record.get(field)) for field in fields)
        if any(values):
            buckets[key_value][values] += 1
    return {item_key: dict(zip(fields, counter.most_common(1)[0][0])) for item_key, counter in buckets.items()}


PROVINCE_NAMES = (
    "内蒙古", "黑龙江", "北京", "天津", "上海", "重庆", "河北", "山西", "辽宁", "吉林",
    "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南", "广东",
    "广西", "海南", "四川", "贵州", "云南", "西藏", "陕西", "甘肃", "青海", "宁夏", "新疆",
    "台湾", "香港", "澳门",
)
MUNICIPALITIES = {"北京", "天津", "上海", "重庆"}
AUTONOMOUS_REGIONS = {"内蒙古": "内蒙古自治区", "广西": "广西壮族自治区", "西藏": "西藏自治区", "宁夏": "宁夏回族自治区", "新疆": "新疆维吾尔自治区"}


def province_from_sales_region(value: Any) -> str:
    """Read a province only when it is explicitly named in the sales-region field."""
    text = clean_str(value).replace(" ", "")
    for name in PROVINCE_NAMES:
        if text.startswith(name):
            if name in MUNICIPALITIES:
                return f"{name}市"
            return AUTONOMOUS_REGIONS.get(name, f"{name}省")
    return ""


def city_aliases(value: Any) -> set[str]:
    """Build conservative aliases for an explicitly supplied city/district name."""
    text = clean_str(value).replace(" ", "")
    if not text:
        return set()
    aliases = {text}
    for province in PROVINCE_NAMES:
        if text.startswith(province) and len(text) > len(province):
            aliases.add(text[len(province):])
    for suffix in ("市区", "地区", "自治州", "盟", "市", "县", "区"):
        for alias in list(aliases):
            if alias.endswith(suffix) and len(alias) > len(suffix):
                aliases.add(alias[:-len(suffix)])
    return {alias for alias in aliases if len(alias) >= 2}


def load_org_customer_supplements(path: Path) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    relationships = pd.read_excel(path, sheet_name="客户对应关系", header=1, engine="openpyxl")
    customers: dict[str, dict[str, str]] = {}
    for row in relationships.to_dict("records"):
        customer = clean_code(row.get("售达方"))
        if not customer:
            continue
        customers[customer] = {
            "channel": clean_str(row.get("渠道")),
            "region": clean_str(row.get("大区")),
            "city": clean_str(row.get("城市")),
        }

    cities = pd.read_excel(path, sheet_name="城市", engine="openpyxl")
    code_to_province: dict[str, str] = {}
    current_province = ""
    for row in cities.to_dict("records"):
        code = clean_code(row.get("SDst"))
        name = clean_str(row.get("地区名字"))
        if not code or not name:
            continue
        if code.endswith("0000") and "省级" in name:
            current_province = province_from_sales_region(name.replace("(省级)", ""))
        if current_province:
            code_to_province[code] = current_province
    alias_votes: dict[str, set[str]] = {}
    for row in cities.to_dict("records"):
        province = code_to_province.get(clean_code(row.get("SDst")), "")
        if not province:
            continue
        for alias in city_aliases(row.get("地区名字")):
            alias_votes.setdefault(alias, set()).add(province)
    # Only retain aliases that identify exactly one province; ambiguous names are not guessed.
    sales_region_to_province = {
        alias: next(iter(provinces))
        for alias, provinces in alias_votes.items()
        if len(provinces) == 1
    }
    return customers, sales_region_to_province


def province_from_city(value: Any, city_to_province: dict[str, str]) -> str:
    provinces = {
        city_to_province[alias]
        for alias in city_aliases(value)
        if alias in city_to_province
    }
    return next(iter(provinces)) if len(provinces) == 1 else ""


def load_reference_html_maps(path: Path) -> dict[str, dict[str, str]]:
    """Read only unambiguous customer-ID mappings explicitly embedded in the reference dashboard."""
    if not path.exists():
        return {}
    _, _, data = extract_app_data(path.read_text(encoding="utf-8"))
    votes: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for row in data.get("records", []):
        customer = clean_code(row.get("cid"))
        if not customer:
            continue
        for field in ("mcid", "mcn", "ch", "p", "ct", "nr"):
            value = clean_str(row.get(field))
            if value and value not in {UNMAPPED, "未识别省份"}:
                votes[customer][field].add(value)
    return {
        customer: {
            field: next(iter(values))
            for field, values in fields.items()
            if len(values) == 1
        }
        for customer, fields in votes.items()
    }


REGION_LABELS = {
    "东北区", "华北区", "西北区", "西南区", "华中区", "华东区", "华南区", "烘焙南区",
    "MKA", "RKA", "出口", "电商", "GKA", "BKA",
}
EXPLICIT_REGION_OVERRIDES = {"7363": "华北区"}


def load_26_01_region_map(path: Path) -> dict[str, str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    mapping: dict[str, str] = {}
    try:
        sheets = [ws for ws in workbook.worksheets if "26年1月" in ws.title]
        if not sheets:
            raise ValueError(f"{path} does not contain the 26年1月 sheet")
        for ws in sheets:
            current_region = ""
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2000), values_only=True):
                raw_region = clean_str(row[1] if len(row) > 1 else None).replace("大区", "区").replace("部门", "")
                if raw_region in REGION_LABELS:
                    current_region = raw_region
                office = clean_str(row[2] if len(row) > 2 else None)
                code = office_code(office)
                if not code or not current_region:
                    continue
                previous = mapping.get(code)
                if previous and previous != current_region:
                    raise ValueError(f"Office {code} has conflicting regions: {previous} / {current_region}")
                mapping[code] = current_region
    finally:
        workbook.close()
    mapping.update(EXPLICIT_REGION_OVERRIDES)
    return mapping


def build_lookup_context(region_master: Path) -> dict[str, Any]:
    profitability_root = REPO_ROOT.parent / "pfs-profitability-dashboard"
    sys.path.insert(0, str(profitability_root))
    from src.pipeline import build_strict_dimension_maps

    canonical_channel, _ = build_strict_dimension_maps(
        F_DIR / ORG_MASTER_NAME,
        F_DIR / CUSTOMER_MASTER_NAME,
        channel_classification_files=[F_DIR / CHANNEL_CLASSIFICATION_NAME, F_DIR / CUSTOMER_LIST_NAME],
    )
    canonical_region = load_26_01_region_map(region_master)
    org_customers, org_sales_region_provinces = load_org_customer_supplements(F_DIR / ORG_MASTER_NAME)
    reference_maps = load_reference_html_maps(F_DIR / REFERENCE_HTML_NAME)
    classification = pd.read_excel(F_DIR / CHANNEL_CLASSIFICATION_NAME, engine="openpyxl")
    canonical_geo: dict[str, dict[str, str]] = {}
    sales_regions: dict[str, str] = {}
    for row in classification.to_dict("records"):
        customer = clean_code(row.get("客户"))
        if not customer:
            continue
        city = clean_str(row.get("Unnamed: 28")) or clean_str(row.get("城市"))
        sales_region = clean_str(row.get("销售地区"))
        province = province_from_sales_region(sales_region) or org_sales_region_provinces.get(sales_region, "")
        geo = canonical_geo.setdefault(customer, {})
        if city and not geo.get("ct"):
            geo["ct"] = city
        if province and not geo.get("p"):
            geo["p"] = province
        if sales_region and customer not in sales_regions:
            sales_regions[customer] = sales_region

    customer_base = pd.read_excel(
        F_DIR / CUSTOMER_MASTER_NAME,
        sheet_name="基础2渠道类型",
        engine="openpyxl",
    )
    customer_names = {
        clean_code(row.get("客户")): clean_str(row.get("客户名称"))
        for row in customer_base.to_dict("records")
        if clean_code(row.get("客户"))
    }
    customer_cities = {
        clean_code(row.get("客户")): clean_str(row.get("城市"))
        for row in customer_base.to_dict("records")
        if clean_code(row.get("客户")) and clean_str(row.get("城市"))
    }
    merge_sheet = pd.read_excel(
        F_DIR / CUSTOMER_MASTER_NAME,
        sheet_name="基础3客户合并",
        engine="openpyxl",
    )
    merged_customers = {
        clean_code(row.get("客户编码")): clean_code(row.get("合并-促销"))
        for row in merge_sheet.to_dict("records")
        if clean_code(row.get("客户编码")) and clean_code(row.get("合并-促销"))
    }
    # Search the whole customer master by exact customer ID. For customers with an
    # explicit promotion merge, the merged customer's city is an allowed fallback.
    for customer, city in customer_cities.items():
        geo = canonical_geo.setdefault(customer, {})
        if city and not geo.get("ct"):
            geo["ct"] = city
        province = province_from_city(city, org_sales_region_provinces)
        if province and not geo.get("p"):
            geo["p"] = province
    customer_list = pd.read_excel(F_DIR / CUSTOMER_LIST_NAME, engine="openpyxl")
    listed_names = {
        clean_code(row.get("客户")): clean_str(row.get("客户编号1"))
        for row in customer_list.to_dict("records")
        if clean_code(row.get("客户"))
    }
    listed_offices = {
        clean_code(row.get("客户")): clean_code(row.get("SOff."))
        for row in customer_list.to_dict("records")
        if clean_code(row.get("客户")) and clean_code(row.get("SOff."))
    }
    return {
        "canonical_channel": canonical_channel,
        "canonical_region": canonical_region,
        "canonical_geo": canonical_geo,
        "customer_names": customer_names,
        "customer_cities": customer_cities,
        "merged_customers": merged_customers,
        "listed_names": listed_names,
        "listed_offices": listed_offices,
        "sales_regions": sales_regions,
        "org_customers": org_customers,
        "reference_maps": reference_maps,
    }


def office_code(value: Any) -> str:
    match = re.search(r"(?<!\d)(\d{4})(?!\d)", clean_str(value))
    return match.group(1) if match else ""


def read_workbook(path: Path, sheet: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet, header=1, engine="openpyxl")


def has_delete_marker(*values: Any) -> bool:
    return any(DELETE_MARKER in clean_str(value) for value in values)


def office_3005_mask(df: pd.DataFrame) -> pd.Series:
    if "列1" in df.columns:
        office = df["列1"].map(clean_code)
    elif "销售办事处" in df.columns:
        office = df["销售办事处"].map(clean_code)
    else:
        return pd.Series(False, index=df.index)
    return office.str.startswith("3005", na=False)


def filter_2026(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    code = df["物料号"].map(clean_code)
    delete_mask = df.apply(lambda row: has_delete_marker(row.get("物料描述"), row.get("品类"), row.get("品项")), axis=1)
    fake_mask = code.isin(BAD_CODES) | delete_mask
    office_mask = office_3005_mask(df)
    kept = df[~fake_mask & ~office_mask].copy()
    return kept, {
        "raw_rows": int(len(df)),
        "kept_gross_income_k": k(kept["销售额"].map(num).sum()),
        "excluded_fake_product_rows": int(fake_mask.sum()),
        "excluded_fake_product_income_k": k(df.loc[fake_mask, "销售额"].map(num).sum()),
        "excluded_fake_product_net_margin_k": k(df.loc[fake_mask, "扣除折让运费净边贡"].map(num).sum()),
        "excluded_fake_product_volume_ton": k(df.loc[fake_mask, "销量KG"].map(num).sum()),
        "excluded_3005_rows": int((~fake_mask & office_mask).sum()),
        "excluded_3005_income_k": k(df.loc[~fake_mask & office_mask, "销售额"].map(num).sum()),
        "excluded_3005_net_margin_k": k(df.loc[~fake_mask & office_mask, "扣除折让运费净边贡"].map(num).sum()),
        "excluded_3005_volume_ton": k(df.loc[~fake_mask & office_mask, "销量KG"].map(num).sum()),
    }


def filter_2025(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    code = df["物料号"].map(clean_code)
    delete_mask = df.apply(lambda row: has_delete_marker(row.get("物料描述"), row.get("品类"), row.get("品项")), axis=1)
    fake_mask = code.isin(BAD_CODES) | delete_mask
    office_mask = office_3005_mask(df)
    kept = df[~fake_mask & ~office_mask].copy()
    return kept, {
        "raw_rows": int(len(df)),
        "excluded_fake_product_rows": int(fake_mask.sum()),
        "excluded_fake_product_income_k": k(df.loc[fake_mask, "销售额"].map(num).sum()),
        "excluded_fake_product_net_margin_k": k(df.loc[fake_mask, "扣除折让运费的净边贡"].map(num).sum()),
        "excluded_fake_product_volume_ton": k(df.loc[fake_mask, "销量KG"].map(num).sum()),
        "excluded_3005_rows": int((~fake_mask & office_mask).sum()),
        "excluded_3005_income_k": k(df.loc[~fake_mask & office_mask, "销售额"].map(num).sum()),
        "excluded_3005_net_margin_k": k(df.loc[~fake_mask & office_mask, "扣除折让运费的净边贡"].map(num).sum()),
        "excluded_3005_volume_ton": k(df.loc[~fake_mask & office_mask, "销量KG"].map(num).sum()),
    }


def enrich_sales(record: dict[str, Any], context: dict[str, Any]) -> dict[str, str]:
    cid = clean_code(record.get("客户编码"))
    cn = clean_str(record.get("客户描述"))
    office = clean_str(record.get("销售办事处"))
    geo = context["canonical_geo"].get(cid) or {}
    org_customer = context["org_customers"].get(cid) or {}
    reference = context["reference_maps"].get(cid) or {}
    researched = RESEARCHED_GEO.get(cid) or {}
    code = office_code(office)
    canonical_region = context["canonical_region"].get(code, "")
    merged_id = context["merged_customers"].get(cid) or reference.get("mcid") or cid
    merged_geo = context["canonical_geo"].get(merged_id) or {}
    merged_org_customer = context["org_customers"].get(merged_id) or {}
    city = (
        geo.get("ct")
        or context["customer_cities"].get(cid)
        or org_customer.get("city")
        or merged_geo.get("ct")
        or context["customer_cities"].get(merged_id)
        or merged_org_customer.get("city")
        or reference.get("ct")
        or researched.get("ct")
        or UNMAPPED
    )
    province = "山东省" if code == "7420" else geo.get("p") or merged_geo.get("p") or reference.get("p") or researched.get("p") or "未识别省份"
    if code == "7420":
        city = "德州"
    region = canonical_region or UNMAPPED
    return {
        "p": province,
        "ct": city,
        "ch": "KA" if code == "7420" else context["canonical_channel"].get(cid) or org_customer.get("channel") or reference.get("ch") or UNMAPPED,
        "nr": region,
        "mcid": merged_id,
        "mcn": context["customer_names"].get(merged_id) or context["listed_names"].get(merged_id) or reference.get("mcn") or (cn if merged_id == cid else UNMAPPED),
    }


def enrich_yoy(customer_id: str, customer_name: str, office: str, context: dict[str, Any]) -> dict[str, str]:
    enriched = enrich_sales({"客户编码": customer_id, "客户描述": customer_name, "销售办事处": office}, context)
    return {
        "p": enriched["p"],
        "ct": enriched["ct"],
        "ch": enriched["ch"],
        "nr": enriched["nr"],
        "mcid": enriched["mcid"],
        "mcn": enriched["mcn"],
    }


def build_sales_records(df: pd.DataFrame, context: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in df.to_dict("records"):
        mo, mr = month_2026(row.get("月份"))
        enriched = enrich_sales(row, context)
        records.append({
            "mo": mo,
            "mr": mr,
            "o": clean_str(row.get("销售办事处")),
            "nr": enriched["nr"],
            "ch": enriched["ch"],
            "cid": clean_code(row.get("客户编码")),
            "cn": clean_str(row.get("客户描述")),
            "mcid": enriched["mcid"],
            "mcn": enriched["mcn"],
            "p": enriched["p"],
            "ct": enriched["ct"],
            "pc": clean_code(row.get("物料号")),
            "pn": clean_str(row.get("物料描述")),
            "cat": clean_str(row.get("品类")),
            "item": clean_str(row.get("品项")),
            "v": k(num(row.get("销量KG"))),
            "inc": k(num(row.get("销售额"))),
            "dt": k(num(row.get("返利"))),
            "am": k(num(row.get("实际出厂边贡"))),
            "fr": k(num(row.get("运费合计"))),
            "ad": k(num(row.get("分摊折让"))),
            "nm": k(num(row.get("扣除折让运费净边贡"))),
            "ni": k(num(row.get("销售收入"))),
        })
    return records


def build_yoy_rows_2025(df: pd.DataFrame, context: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in df.to_dict("records"):
        customer_id = clean_code(row.get("客户编码"))
        customer_name = clean_str(row.get("客户描述"))
        office = clean_str(row.get("销售办事处"))
        enriched = enrich_yoy(customer_id, customer_name, office, context)
        records.append({
            "y": "2025",
            "mo": month_cn(row.get("月份"), "2025"),
            "cat": clean_str(row.get("品类")),
            "item": clean_str(row.get("品项")),
            "ch": enriched["ch"],
            "nr": enriched["nr"],
            "p": enriched["p"],
            "ct": enriched["ct"],
            "mcid": enriched["mcid"],
            "mcn": enriched["mcn"],
            "pc": clean_code(row.get("物料号")),
            "pn": clean_str(row.get("物料描述")),
            "o": office,
            "src": 1,
            "v": k(num(row.get("销量KG"))),
            "inc": k(num(row.get("销售额"))),
            "ad": k(num(row.get("分摊的折让"))),
            "nm": k(num(row.get("扣除折让运费的净边贡"))),
            "ni": k(num(row.get("销售收入"))),
        })
    return records


def build_yoy_rows_2026(df: pd.DataFrame, context: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in df.to_dict("records"):
        customer_id = clean_code(row.get("客户编码"))
        customer_name = clean_str(row.get("客户描述"))
        office = clean_str(row.get("销售办事处"))
        enriched = enrich_yoy(customer_id, customer_name, office, context)
        records.append({
            "y": "2026",
            "mo": month_cn(row.get("月份"), "2026"),
            "cat": clean_str(row.get("品类")),
            "item": clean_str(row.get("品项")),
            "ch": enriched["ch"],
            "nr": enriched["nr"],
            "p": enriched["p"],
            "ct": enriched["ct"],
            "mcid": enriched["mcid"],
            "mcn": enriched["mcn"],
            "pc": clean_code(row.get("物料号")),
            "pn": clean_str(row.get("物料描述")),
            "o": office,
            "src": 1,
            "v": k(num(row.get("销量KG"))),
            "inc": k(num(row.get("销售额"))),
            "ad": k(num(row.get("分摊折让"))),
            "nm": k(num(row.get("扣除折让运费净边贡"))),
            "ni": k(num(row.get("销售收入"))),
        })
    return records


def aggregate_yoy(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    dims = ("y", "mo", "cat", "item", "ch", "nr", "p", "ct", "mcid", "mcn", "pc", "pn", "o")
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for record in records:
        key = tuple(record[dim] for dim in dims)
        if key not in grouped:
            grouped[key] = {dim: record[dim] for dim in dims}
            grouped[key].update({"src": 0, "v": 0.0, "inc": 0.0, "ad": 0.0, "nm": 0.0, "ni": 0.0})
        target = grouped[key]
        target["src"] += int(record.get("src") or 1)
        target["v"] += float(record.get("v") or 0)
        target["inc"] += float(record.get("inc") or 0)
        target["ad"] += float(record.get("ad") or 0)
        target["nm"] += float(record.get("nm") or 0)
        target["ni"] += float(record.get("ni") or 0)
    result = list(grouped.values())
    for record in result:
        record["v"] = round(record["v"], 6)
        record["inc"] = round(record["inc"], 6)
        record["ad"] = round(record["ad"], 6)
        record["nm"] = round(record["nm"], 6)
        record["ni"] = round(record["ni"], 6)
    result.sort(key=lambda item: (item["y"], item["mo"], item["cat"], item["item"], item["ch"], item["p"], item["ct"], item["mcid"], item["pc"], item["o"]))
    return result


def totals(records: list[dict[str, Any]]) -> dict[str, float]:
    volume = sum(float(record.get("v") or 0) for record in records)
    income = sum(float(record.get("inc") or 0) for record in records)
    net_income = sum(float(record.get("ni") or 0) for record in records)
    allocated_discount = sum(float(record.get("ad") or 0) for record in records)
    net_margin = sum(float(record.get("nm") or 0) for record in records)
    net_margin_base = net_income - allocated_discount
    return {
        "volume_ton": round(volume, 6),
        "income_k": round(income, 6),
        "net_income_k": round(net_income, 6),
        "allocated_discount_k": round(allocated_discount, 6),
        "net_margin_base_k": round(net_margin_base, 6),
        "net_margin_k": round(net_margin, 6),
        "net_margin_rate": net_margin / net_margin_base if net_margin_base else 0.0,
    }


def build_unmapped_rows(df: pd.DataFrame, year: int, context: dict[str, Any]) -> list[dict[str, str]]:
    output: dict[tuple[str, ...], dict[str, str]] = {}
    for row in df.to_dict("records"):
        customer_id = clean_code(row.get("客户编码"))
        customer_name = clean_str(row.get("客户描述"))
        office = clean_str(row.get("销售办事处"))
        code = office_code(office)
        enriched = enrich_sales(row, context)
        missing: list[str] = []
        if enriched["ch"] == UNMAPPED:
            missing.append("渠道")
        if customer_id in context["merged_customers"] and enriched["mcn"] == UNMAPPED:
            missing.append("合并客户名称")
        if enriched["ct"] == UNMAPPED:
            missing.append("城市")
        if enriched["p"] in {UNMAPPED, "未识别省份"}:
            missing.append("省份")
        if enriched["nr"] == UNMAPPED:
            missing.append("大区")
        if not missing:
            continue
        raw_month = clean_code(row.get("月份"))
        month = int(raw_month[-2:]) if raw_month else 0
        key = (str(year), str(month), customer_id, customer_name, office, "、".join(missing))
        output[key] = {
            "年份": str(year),
            "来源月份": f"{year}-{month:02d}",
            "客户编码": customer_id,
            "客户名称": customer_name,
            "销售办事处": office,
            "销售地区": context["sales_regions"].get(customer_id, ""),
            "缺失字段": "、".join(missing),
        }
    return list(output.values())


def build_region_unmapped_rows(
    records: list[dict[str, Any]],
    source_report: str,
    region_map: dict[str, str],
    year_field: str,
    sales_field: str,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for record in records:
        office = clean_str(record.get("o"))
        code = office_code(office)
        if code in region_map:
            continue
        year = clean_str(record.get(year_field))
        key = (source_report, year, code or "未提取", office)
        row = grouped.setdefault(
            key,
            {
                "来源报表": source_report,
                "年份": year,
                "销售办公室编码": code or "未提取",
                "销售办公室": office,
                "记录数": 0,
                "销售额": 0.0,
            },
        )
        row["记录数"] += 1
        row["销售额"] += float(record.get(sales_field) or 0)
    return list(grouped.values())


def write_unmapped_report(
    rows: list[dict[str, str]],
    region_rows: list[dict[str, Any]] | None = None,
) -> None:
    columns = ["年份", "来源月份", "客户编码", "客户名称", "销售办事处", "销售地区", "缺失字段"]
    detail = pd.DataFrame(rows, columns=columns).sort_values(columns, kind="stable")
    summary = (
        detail.assign(缺失字段=detail["缺失字段"].str.split("、"))
        .explode("缺失字段")
        .groupby(["年份", "缺失字段"], dropna=False)
        .size()
        .reset_index(name="记录数")
    )
    with pd.ExcelWriter(UNMAPPED_OUTPUT, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="未映射明细", index=False)
        summary.to_excel(writer, sheet_name="汇总", index=False)
        pd.DataFrame(
            region_rows or [],
            columns=["来源报表", "年份", "销售办公室编码", "销售办公室", "记录数", "销售额"],
        ).sort_values(
            ["来源报表", "年份", "销售办公室编码"], kind="stable"
        ).to_excel(writer, sheet_name="大区未映射", index=False)


def update_quality_sales(data: dict[str, Any], records: list[dict[str, Any]], source_2026: Path, stats_2026: dict[str, Any]) -> None:
    quality = data.setdefault("quality", {})
    quality["period"] = "2026.01-2026.07"
    quality["source"] = str(source_2026)
    quality["income_metric_source"] = "销售额（含税销额）"
    quality["gross_income_metric_source"] = "销售额（含税销额）"
    quality["net_margin_rate_formula"] = "扣除折让运费后的净边贡 / (销售收入 - 分摊后折让)"
    quality["net_income_metric_source"] = "销售收入（未税收入）"
    quality["rows"] = len(records)
    quality["month_order"] = ["2026.01", "2026.02", "2026.03", "2026.04", "2026.05", "2026.06", "2026.07"]
    quality["sku_cnt"] = len({record["pc"] for record in records if record.get("pc")})
    quality["customer_cnt"] = len({record["cid"] for record in records if record.get("cid")})
    quality["merged_customer_cnt"] = len({record["mcid"] for record in records if record.get("mcid")})
    quality["category_cnt"] = len({record["cat"] for record in records if record.get("cat")})
    quality["item_cnt"] = len({record["item"] for record in records if record.get("item")})
    quality["channel_cnt"] = len({record["ch"] for record in records if record.get("ch")})
    total_volume = sum(float(record.get("v") or 0) for record in records)
    total_income = sum(float(record.get("inc") or 0) for record in records)
    total_net_income = sum(float(record.get("ni") or 0) for record in records)
    total_allocated_discount = sum(float(record.get("ad") or 0) for record in records)
    total_net_margin = sum(float(record.get("nm") or 0) for record in records)
    total_net_margin_base = total_net_income - total_allocated_discount
    quality["total_volume_ton"] = round(total_volume, 6)
    quality["total_gross_income_k"] = stats_2026["kept_gross_income_k"]
    quality["total_income_k"] = round(total_income, 6)
    quality["total_net_income_k"] = round(total_net_income, 6)
    quality["total_allocated_discount_k"] = round(total_allocated_discount, 6)
    quality["total_net_margin_base_k"] = round(total_net_margin_base, 6)
    quality["total_net_margin_k"] = round(total_net_margin, 6)
    quality["net_margin_rate"] = total_net_margin / total_net_margin_base if total_net_margin_base else 0.0
    quality["excluded_fake_product_rows"] = stats_2026["excluded_fake_product_rows"]
    quality["excluded_fake_product_income_k"] = stats_2026["excluded_fake_product_income_k"]
    quality["excluded_fake_product_volume_ton"] = stats_2026["excluded_fake_product_volume_ton"]
    quality["excluded_fake_product_net_margin_k"] = stats_2026["excluded_fake_product_net_margin_k"]
    quality["excluded_3005_rows"] = stats_2026["excluded_3005_rows"]
    quality["excluded_3005_income_k"] = stats_2026["excluded_3005_income_k"]
    quality["excluded_3005_volume_ton"] = stats_2026["excluded_3005_volume_ton"]
    quality["excluded_3005_net_margin_k"] = stats_2026["excluded_3005_net_margin_k"]
    quality["missing_customer_cnt"] = len({record["cid"] for record in records if record.get("p") == "未识别"})


def update_quality_yoy(data: dict[str, Any], records: list[dict[str, Any]], source_2025: Path, source_2026: Path, stats_2025: dict[str, Any], stats_2026: dict[str, Any]) -> None:
    quality = data.setdefault("quality", {})
    quality["source_2025"] = str(source_2025)
    quality["source_2026"] = str(source_2026)
    quality["income_metric_source_2026"] = "销售额（含税销额）"
    quality["income_metric_source_2025"] = "销售额（含税销额）"
    quality["gross_income_metric_source_2026"] = "销售额（含税销额）"
    quality["gross_income_metric_source_2025"] = "销售额（含税销额）"
    quality["net_margin_rate_formula"] = "扣除折让运费后的净边贡 / (销售收入 - 分摊后折让)"
    column_map = quality.setdefault("metric_column_map", {})
    column_map.setdefault("2026", {})["income"] = "销售额 / 1000（含税销额）"
    column_map.setdefault("2026", {})["gross_income"] = "销售额 / 1000（含税销额）"
    column_map.setdefault("2026", {})["net_income"] = "销售收入 / 1000（未税收入，用于净边贡率分母）"
    column_map.setdefault("2025", {})["income"] = "销售额 / 1000（含税销额）"
    column_map.setdefault("2025", {})["gross_income"] = "销售额 / 1000（含税销额）"
    column_map.setdefault("2025", {})["net_income"] = "销售收入 / 1000（未税收入，用于净边贡率分母）"
    quality["aggregated_rows"] = len(records)
    quality["category_cnt"] = len({record["cat"] for record in records if record.get("cat")})
    quality["totals"] = {
        "2025": totals([record for record in records if record.get("y") == "2025"]),
        "2026": totals([record for record in records if record.get("y") == "2026"]),
    }
    quality["raw_rows_2025"] = stats_2025["raw_rows"]
    quality["excluded_fake_product_rows_2025"] = stats_2025["excluded_fake_product_rows"]
    quality["excluded_fake_product_income_k_2025"] = stats_2025["excluded_fake_product_income_k"]
    quality["excluded_fake_product_net_margin_k_2025"] = stats_2025["excluded_fake_product_net_margin_k"]
    quality["excluded_fake_product_volume_ton_2025"] = stats_2025["excluded_fake_product_volume_ton"]
    quality["excluded_3005_rows_2025"] = stats_2025["excluded_3005_rows"]
    quality["excluded_3005_income_k_2025"] = stats_2025["excluded_3005_income_k"]
    quality["excluded_3005_net_margin_k_2025"] = stats_2025["excluded_3005_net_margin_k"]
    quality["excluded_3005_volume_ton_2025"] = stats_2025["excluded_3005_volume_ton"]
    quality["raw_rows_2026"] = stats_2026["raw_rows"]
    quality["excluded_fake_product_rows_2026"] = stats_2026["excluded_fake_product_rows"]
    quality["excluded_fake_product_income_k_2026"] = stats_2026["excluded_fake_product_income_k"]
    quality["excluded_fake_product_net_margin_k_2026"] = stats_2026["excluded_fake_product_net_margin_k"]
    quality["excluded_fake_product_volume_ton_2026"] = stats_2026["excluded_fake_product_volume_ton"]
    quality["excluded_3005_rows_2026"] = stats_2026["excluded_3005_rows"]
    quality["excluded_3005_income_k_2026"] = stats_2026["excluded_3005_income_k"]
    quality["excluded_3005_net_margin_k_2026"] = stats_2026["excluded_3005_net_margin_k"]
    quality["excluded_3005_volume_ton_2026"] = stats_2026["excluded_3005_volume_ton"]
    quality["excluded_3005_rows"] = stats_2026["excluded_3005_rows"]
    quality["excluded_3005_income_k"] = stats_2026["excluded_3005_income_k"]
    quality["excluded_3005_net_margin_k"] = stats_2026["excluded_3005_net_margin_k"]
    quality["excluded_3005_volume_ton"] = stats_2026["excluded_3005_volume_ton"]


def main() -> None:
    parser = argparse.ArgumentParser(description="按26年1月销售办公室主数据生成TBFX看板")
    parser.add_argument("--region-master", required=True, type=Path, help="包含26年1月Sheet的销售办公室主数据")
    args = parser.parse_args()
    source_2026, source_2025 = copy_sources_to_f_dir()
    sales_data = load_app_data(SALES_STATIC)
    yoy_data = load_app_data(YOY_STATIC)
    context = build_lookup_context(args.region_master)

    df_2026 = read_workbook(source_2026, "26")
    df_2025 = read_workbook(source_2025, "25")
    clean_2026, stats_2026 = filter_2026(df_2026)
    clean_2025, stats_2025 = filter_2025(df_2025)

    sales_records = build_sales_records(clean_2026, context)
    yoy_raw = build_yoy_rows_2025(clean_2025, context) + build_yoy_rows_2026(clean_2026, context)
    yoy_records = aggregate_yoy(yoy_raw)

    unmapped_rows = build_unmapped_rows(clean_2025, 2025, context) + build_unmapped_rows(clean_2026, 2026, context)
    region_unmapped_rows = build_region_unmapped_rows(
        sales_records,
        "2026销售报表",
        context["canonical_region"],
        "mo",
        "inc",
    ) + build_region_unmapped_rows(
        yoy_records,
        "2025/2026同比报表",
        context["canonical_region"],
        "y",
        "inc",
    )
    write_unmapped_report(unmapped_rows, region_unmapped_rows)

    sales_data["records"] = sales_records
    sales_data["monthOrder"] = ["2026.01", "2026.02", "2026.03", "2026.04", "2026.05", "2026.06", "2026.07"]
    sales_data["channelOrder"] = YOY_CHANNEL_ORDER
    update_quality_sales(sales_data, sales_records, source_2026, stats_2026)
    yoy_data["records"] = yoy_records
    yoy_data["monthOrder"] = YOY_MONTH_ORDER
    yoy_data["channelOrder"] = YOY_CHANNEL_ORDER
    update_quality_yoy(yoy_data, yoy_records, source_2025, source_2026, stats_2025, stats_2026)

    for path, data in ((SALES_STATIC, sales_data), (YOY_STATIC, yoy_data)):
        replace_app_data(path, data)

    shutil.copy2(SALES_STATIC, SALES_SOURCE_HTML)
    shutil.copy2(YOY_STATIC, YOY_SOURCE_HTML)

    print(json.dumps({
        "source_2026": str(source_2026),
        "source_2025": str(source_2025),
        "sales_records": len(sales_records),
        "yoy_records": len(yoy_records),
        "sales_totals": totals(sales_records),
        "yoy_totals": yoy_data["quality"]["totals"],
        "stats_2026": stats_2026,
        "stats_2025": stats_2025,
        "unmapped_report": str(UNMAPPED_OUTPUT),
        "unmapped_records": len(unmapped_rows),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
